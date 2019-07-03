from openpyxl import load_workbook
from flask import Flask, request, jsonify

app = Flask(__name__)

data = load_workbook(filename='2019-1a.xlsx')['Sheet1']
# print(data.cell(row=1, column=1).value)

keys = []
for cell in data[1]:
    keys.append(cell.value)

def query(name):
    result = []
    for row in data.iter_rows():
        if row[0].value == name:
            for cell in row:
                result.append(cell.value)
    return dict(zip(keys, result))

@app.route('/api/get/<name>', methods=['GET'])
def get(name):
    print(name)
    return jsonify(query(name))